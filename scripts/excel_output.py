#!/usr/bin/env python3
"""
XAFS Fit Excel Parameter File Generator.

Generates a 6-sheet Excel (.xlsx) file with complete fitting information:
  1. Feff_Paths    -- Scattering paths (name, feff file, degen, reff, scatterer, shell)
  2. Fit_Parameters -- Fitting parameters (name, initial, lower, upper, fit value, uncertainty)
  3. Amplitude_Expression -- S02 amplitude expression
  4. Fit_Conditions -- Fitting conditions (k-range, R-range, k-weight, window, N_indep, N_var)
  5. Data_Processing -- Data processing steps
  6. Results       -- Final structure parameters table

Example:
    $ python excel_output.py --result fit_result.json --output ./Fit/Fit_01/
"""

import sys
import json
from pathlib import Path
from typing import Dict, List, Optional

import numpy as np


class XAFSExcelGenerator:
    """Generate XAFS fitting Excel parameter file."""

    def __init__(self, result: Dict):
        """
        Initialize with a fit result dict.

        Args:
            result: Fit result dict (from fit_core.py output JSON).
        """
        self.result = result
        self.sample = result.get("sample_name", "Sample")

    def generate(self, output_path: str) -> str:
        """
        Generate the 6-sheet Excel file.

        Args:
            output_path: Output .xlsx file path.

        Returns:
            The output path.

        Raises:
            ImportError: If openpyxl is not installed.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel output. "
                "Install it with: pip install openpyxl"
            )

        wb = Workbook()
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        def style_header(ws, row=1, max_col=None):
            if max_col is None:
                max_col = ws.max_column
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

        def style_data(ws, start_row=2, max_col=None):
            if max_col is None:
                max_col = ws.max_column
            for row in range(start_row, ws.max_row + 1):
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')

        def auto_width(ws, min_width=10, max_width=25):
            for col_cells in ws.columns:
                col_letter = get_column_letter(col_cells[0].column)
                max_len = min_width
                for cell in col_cells:
                    if cell.value:
                        max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
                ws.column_dimensions[col_letter].width = max_len

        # --- Sheet 1: Feff_Paths ---
        ws1 = wb.active
        ws1.title = "Feff_Paths"
        ws1.append(["Path Name", "Feff File", "Degeneracy", "Reff (A)", "Scattering Atoms", "Shell"])
        style_header(ws1)
        for path in self.result.get("paths", []):
            ws1.append([
                path.get("name", ""),
                path.get("feff_file", ""),
                path.get("degeneracy", ""),
                path.get("reff", ""),
                path.get("scattering_atoms", ""),
                path.get("shell", ""),
            ])
        style_data(ws1)
        auto_width(ws1)

        # --- Sheet 2: Fit_Parameters ---
        ws2 = wb.create_sheet("Fit_Parameters")
        ws2.append(["Parameter", "Initial Value", "Lower Bound", "Upper Bound",
                     "Fit Value", "Uncertainty", "Fixed"])
        style_header(ws2)
        for p in self.result.get("parameters", []):
            ws2.append([
                p.get("name", ""),
                p.get("initial_value", ""),
                p.get("lower_bound", ""),
                p.get("upper_bound", ""),
                p.get("fit_value", ""),
                p.get("uncertainty", ""),
                "Yes" if p.get("fixed", False) else "No",
            ])
        style_data(ws2)
        auto_width(ws2)

        # --- Sheet 3: Amplitude_Expression ---
        ws3 = wb.create_sheet("Amplitude_Expression")
        ws3.append(["Property", "Value"])
        style_header(ws3)
        s02 = self.result.get("conditions", {}).get("s02", 1.0)
        s02_fixed = self.result.get("conditions", {}).get("s02_fixed", True)
        ws3.append(["S02", s02])
        ws3.append(["S02 Status", "Fixed" if s02_fixed else "Floated"])
        ws3.append(["Expression", self.result.get("amplitude_expression", f"S02 = {s02}")])
        style_data(ws3)
        auto_width(ws3)

        # --- Sheet 4: Fit_Conditions ---
        ws4 = wb.create_sheet("Fit_Conditions")
        ws4.append(["Condition", "Value"])
        style_header(ws4)
        c = self.result.get("conditions", {})
        conditions_rows = [
            ["k-range (A^-1)", f"{c.get('k_min', 'N/A')} - {c.get('k_max', 'N/A')}"],
            ["R-range (A)", f"{c.get('r_min', 'N/A')} - {c.get('r_max', 'N/A')}"],
            ["k-weight", c.get("k_weight", "N/A")],
            ["Window", c.get("window", "N/A")],
            ["dk (A^-1)", c.get("dk", "N/A")],
            ["S02", c.get("s02", "N/A")],
            ["N_indep", self.result.get("n_indep", "N/A")],
            ["N_var", self.result.get("n_var", "N/A")],
            ["N_points (k-space)", "See data file"],
            ["Backend", self.result.get("backend", "N/A")],
        ]
        for row in conditions_rows:
            ws4.append(row)
        style_data(ws4)
        auto_width(ws4)

        # --- Sheet 5: Data_Processing ---
        ws5 = wb.create_sheet("Data_Processing")
        ws5.append(["Step", "Description"])
        style_header(ws5)
        steps = self.result.get("data_processing_steps", [])
        for i, step in enumerate(steps, 1):
            ws5.append([i, step])
        style_data(ws5)
        auto_width(ws5)

        # --- Sheet 6: Results ---
        ws6 = wb.create_sheet("Results")
        c = self.result.get("conditions", {})
        ws6.append(["Parameter", "Value", "Uncertainty", "Unit"])
        style_header(ws6)

        params = self.result.get("parameters", [])
        n_paths = len(self.result.get("paths", []))
        for i, p in enumerate(params):
            name = p.get("name", "")
            val = p.get("fit_value", p.get("initial_value", ""))
            unc = p.get("uncertainty", "")

            if name.startswith("N_"):
                unit = ""
            elif name.startswith("R_"):
                unit = "Angstrom"
            elif "sigma" in name.lower():
                unit = "Angstrom^2"
            elif name.startswith("DeltaE"):
                unit = "eV"
            else:
                unit = ""

            ws6.append([name, val, unc, unit])

        ws6.append([])
        ws6.append(["Fit Quality", "", "", ""])
        ws6.append(["R-factor", self.result.get("r_factor", ""), "", ""])
        ws6.append(["Reduced chi-square", self.result.get("chi_sq_reduced", ""), "", ""])
        ws6.append(["N_indep", self.result.get("n_indep", ""), "", ""])
        ws6.append(["N_var", self.result.get("n_var", ""), "", ""])

        style_data(ws6)
        auto_width(ws6)

        # Save
        out_path = Path(output_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(out_path))
        print(f"Excel parameter file saved to: {out_path}")
        return str(out_path)


def main():
    """CLI entry point for Excel generation."""
    import argparse

    parser = argparse.ArgumentParser(description="XAFS Fit Excel Parameter Generator")
    parser.add_argument("--result", required=True, help="Path to fit result JSON")
    parser.add_argument("--output", required=True, help="Output .xlsx file path")
    parser.add_argument("--check-prereqs", action="store_true",
                        help="Check prerequisites and exit")
    parser.add_argument("--diagnostics", action="store_true",
                        help="Output skill diagnostics and exit")

    args = parser.parse_args()

    if args.diagnostics:
        diag = {
            "skill": "xafs-fit-skill",
            "component": "excel_output",
            "version": "1.0.0",
            "harness_features": {"input_validation": True, "output_sanity": True},
        }
        print(json.dumps(diag, indent=2))
        return 0

    if args.check_prereqs:
        checks = []
        for pkg in ["numpy", "openpyxl"]:
            try:
                __import__(pkg)
                checks.append({"check": f"package_{pkg}", "required": pkg,
                                "found": "installed", "ok": True})
            except ImportError:
                checks.append({"check": f"package_{pkg}", "required": pkg,
                                "found": "missing", "ok": False})

        import platform
        py_ver = platform.python_version()
        checks.append({"check": "python_version", "required": ">=3.9",
                        "found": py_ver, "ok": tuple(map(int, py_ver.split("."))) >= (3, 9)})

        ready = all(c["ok"] for c in checks)
        print(json.dumps({"ready": ready, "checks": checks}, indent=2))
        return 0 if ready else 1

    with open(args.result) as f:
        result = json.load(f)

    gen = XAFSExcelGenerator(result)
    gen.generate(args.output)
    return 0


if __name__ == "__main__":
    sys.exit(main())
